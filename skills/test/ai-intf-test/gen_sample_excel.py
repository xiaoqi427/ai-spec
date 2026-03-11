#!/usr/bin/env python3
"""
生成示例输入 Excel 文件

根据 BPM 实现清单，生成包含所有继承 BaseClaimWebApi 的 Controller
对应 itemId 的测试输入模板。用户只需填入实际的 claimId 即可使用。

用法:
  python gen_sample_excel.py
  python gen_sample_excel.py --output ./excel/test_input.xlsx
"""
import argparse
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ============================================
# 所有继承 BaseClaimWebApi 的 Controller 清单
# 来源: BPM实现清单.md
# ============================================
CONTROLLERS = [
    ("T001", "差旅费报销单", "eer"),
    ("T002", "员工借款单", "eer"),
    ("T003", "员工还款单", "eer"),
    ("T004", "延迟还款申请单", "eer"),
    ("T005", "备用金及手机额度申请单", "eer"),
    ("T006", "备用金借款单", "eer"),
    ("T007", "对员工支付费用报销单", "eer"),
    ("T008", "对公支付费用报销单", "eer"),
    ("T009", "对公支付营销费用报销单", "eer"),
    ("T010", "调整类业务报账单", "ptp"),
    ("T011", "费用预付款申请单", "eer"),
    ("T012", "费用类计提报账单", "ptp"),
    ("T013", "计提冲销报账单", "eer"),
    ("T014", "承兑汇票总账报账单", "tr"),
    ("T015", "内部资金调拨申请单", "tr"),
    ("T016", "资金其他付款单", "tr"),
    ("T017", "总账收款单", "rtr"),
    ("T018", "总账付款单", "rtr"),
    ("T019", "总账通用单", "rtr"),
    ("T020", "总账薪酬单", "rtr"),
    ("T022", "冻结解冻申请", "base"),
    ("T023", "付款条件变更申请单", "base"),
    ("T026", "采购类预付款申请单", "ptp"),
    ("T027", "物流费用报销报账单", "ptp"),
    ("T028", "总账调账单", "rtr"),
    ("T030", "资产报账单", "fa"),
    ("T031", "预付类通用调账报账单", "base"),
    ("T032", "应付类通用调账报账单", "base"),
    ("T033", "合并付款申请报账单", "base"),
    ("T034", "协同报账单", "ptp"),
    ("T035", "应付薪酬单", "rtr"),
    ("T038", "金德瑞协同报账单", "ptp"),
    ("T039", "供应商协同报账单", "ptp"),
    ("T040", "其他收入收款类报账单", "otc"),
    ("T041", "采购方调拨报账单", "ptp"),
    ("T043", "收款报账单", "otc"),
    ("T044", "承兑汇票开票申请单", "tr"),
    ("T045", "收入成本报账单", "otc"),
    ("T046", "转款报账单", "otc"),
    ("T047", "手工发票报账单", "otc"),
    ("T048", "退款报账单", "otc"),
    ("T049", "杂项采购业务报账单", "ptp"),
    ("T050", "付款对象变更申请单", "base"),
    ("T051", "内部往来报账单", "base"),
    ("T052", "内部往来报账单-接收方", "base"),
    ("T053", "资产报废报账单", "fa"),
    ("T054", "资产调拨（调入）报账单", "fa"),
    ("T056", "资产出售报账单", "fa"),
    ("T057", "外币业务报账单", "base"),
    ("T065", "受托支付报账单（事业部）", "tr"),
    ("T900", "虚拟报账单", "base"),
]


def generate_sample_excel(output_path: str):
    """生成示例输入 Excel"""
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 样式定义
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    note_font = Font(name="微软雅黑", italic=True, size=9, color="808080")

    # 表头
    headers = [
        ("claimId", 18, "报账单ID（必填，整数）"),
        ("itemId", 12, "模板ID（必填，如 T001）"),
        ("description", 30, "说明（可选）"),
        ("module", 10, "所属模块（参考）"),
    ]

    for col_idx, (name, width, _) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 写入示例数据（每个 Controller 一行，claimId 留空供用户填写）
    for row_idx, (item_id, desc, module) in enumerate(CONTROLLERS, start=2):
        # claimId 留空 - 用户需要填入实际的 claimId
        ws.cell(row=row_idx, column=1, value=None).alignment = cell_align
        ws.cell(row=row_idx, column=2, value=item_id).alignment = cell_align
        ws.cell(row=row_idx, column=3, value=desc).alignment = Alignment(vertical="center")
        ws.cell(row=row_idx, column=4, value=module).alignment = cell_align

        for col_idx in range(1, 5):
            ws.cell(row=row_idx, column=col_idx).border = thin_border

    # 添加说明 Sheet
    ws_note = wb.create_sheet(title="使用说明")
    notes = [
        "submitValidate 批量测试 - 输入模板说明",
        "",
        "1. 在 Sheet1 的 claimId 列填入要测试的报账单ID（整数）",
        "2. itemId 列已预填所有继承 BaseClaimWebApi 的模板编号",
        "3. 如果只需测试部分模板，删除不需要的行即可",
        "4. 同一个 itemId 可以填多行不同的 claimId",
        "5. description 和 module 列仅供参考，不影响测试",
        "",
        "运行方式:",
        "  cd ai-intf-test",
        "  pip install -r requirements.txt",
        "  python -m ai_intf_test.main --config config/config.yaml",
        "",
        "或指定输入文件:",
        "  python -m ai_intf_test.main -i ./excel/test_input.xlsx -u http://your-server:8080",
    ]
    for i, line in enumerate(notes, start=1):
        cell = ws_note.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14)
        else:
            cell.font = Font(size=11)
    ws_note.column_dimensions["A"].width = 80

    wb.save(output_path)
    print(f"示例 Excel 已生成: {output_path}")
    print(f"包含 {len(CONTROLLERS)} 个模板，请在 claimId 列填入实际数据后使用")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="生成示例输入 Excel")
    parser.add_argument(
        "--output", "-o",
        type=str,
        default="./excel/test_input.xlsx",
        help="输出文件路径（默认: ./excel/test_input.xlsx）",
    )
    args = parser.parse_args()
    generate_sample_excel(args.output)
