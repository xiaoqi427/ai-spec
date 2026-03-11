"""Excel 读写模块：读取测试输入、写出测试结果"""
from __future__ import annotations

import json
import logging
import os
from datetime import datetime
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import TestInput, TestResult

logger = logging.getLogger(__name__)


def _find_column_index(header_row, column_name: str) -> int:
    """根据列名在表头中找到列索引（1-based），找不到则抛出 KeyError"""
    normalized = column_name.strip().lower()
    for cell in header_row:
        if isinstance(cell.value, str) and cell.value.strip().lower() == normalized:
            return cell.column
    raise KeyError(f"在 Excel 表头中未找到列名: {column_name}")


def read_test_inputs(
    excel_path: str,
    sheet_name: str,
    claim_id_column: str,
    item_id_column: str,
) -> List[TestInput]:
    """从 Excel 读取测试输入数据

    Args:
        excel_path: Excel 文件路径
        sheet_name: 工作表名称
        claim_id_column: claimId 列名
        item_id_column: itemId 列名

    Returns:
        TestInput 列表
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"输入 Excel 文件不存在: {excel_path}")

    wb = load_workbook(excel_path, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Excel 中不存在工作表: {sheet_name}，可用: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows())
    if not rows:
        logger.warning("Excel 工作表为空")
        return []

    header = rows[0]
    claim_id_idx = _find_column_index(header, claim_id_column)
    item_id_idx = _find_column_index(header, item_id_column)

    inputs: List[TestInput] = []
    for row_num, row in enumerate(rows[1:], start=2):
        claim_id_val = row[claim_id_idx - 1].value
        item_id_val = row[item_id_idx - 1].value

        if item_id_val is None or str(item_id_val).strip() == "":
            logger.debug(f"跳过第 {row_num} 行：itemId 为空")
            continue

        if claim_id_val is None or str(claim_id_val).strip() == "":
            logger.debug(f"跳过第 {row_num} 行：claimId 为空")
            continue

        try:
            claim_id = int(claim_id_val)
        except (ValueError, TypeError):
            logger.warning(f"第 {row_num} 行 claimId 无法转为整数: {claim_id_val}，跳过")
            continue

        item_id = str(item_id_val).strip().upper()
        inputs.append(TestInput(claim_id=claim_id, item_id=item_id, row_index=row_num))

    wb.close()
    logger.info(f"从 Excel 读取到 {len(inputs)} 条测试数据")
    return inputs


# ============================
# 输出 Excel 样式常量
# ============================
_HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_ERROR_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# 输出列定义
_OUTPUT_COLUMNS = [
    ("claimId", 15),
    ("itemId", 10),
    ("serviceType", 12),
    ("url", 50),
    ("statusCode", 12),
    ("success", 10),
    ("pass", 10),
    ("results", 80),
    ("error", 40),
    ("traceId", 18),
    ("testTime", 22),
    ("responseTimeMs", 16),
]


def write_test_results(
    results: List[TestResult],
    output_path: str,
) -> str:
    """将测试结果写入 Excel 文件

    Args:
        results: 测试结果列表
        output_path: 输出文件路径（支持 {timestamp} 占位符）

    Returns:
        实际输出的文件路径
    """
    # 替换时间戳占位符
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    actual_path = output_path.replace("{timestamp}", timestamp)

    # 确保输出目录存在
    output_dir = os.path.dirname(actual_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "测试结果"

    # 写入表头
    for col_idx, (col_name, col_width) in enumerate(_OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 写入数据
    for row_idx, result in enumerate(results, start=2):
        # 校验结果详情：转为 JSON 摘要
        results_text = _format_results_text(result)

        row_data = [
            result.claim_id,
            result.item_id,
            result.service_type,
            result.url,
            result.status_code,
            result.success,
            result.passed,
            results_text,
            result.error or "",
            result.trace_id or "",
            result.test_time,
            result.response_time_ms,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = _CELL_ALIGNMENT
            cell.border = _THIN_BORDER

        # 根据结果设置行颜色
        _apply_row_color(ws, row_idx, result, len(_OUTPUT_COLUMNS))

    # 写入统计摘要 Sheet
    _write_summary_sheet(wb, results)

    wb.save(actual_path)
    logger.info(f"测试结果已写入: {actual_path}")
    return actual_path


def _format_results_text(result: TestResult) -> str:
    """格式化校验结果为可读文本"""
    if not result.results:
        return ""
    parts = []
    for r in result.results:
        severity_tag = f"[{r.severity}]"
        parts.append(f"{severity_tag} {r.name}: {r.message}")
    return "\n".join(parts)


def _apply_row_color(ws, row_idx: int, result: TestResult, col_count: int):
    """根据测试结果给行设置背景色"""
    if result.error:
        fill = _ERROR_FILL
    elif result.passed:
        fill = _PASS_FILL
    else:
        fill = _FAIL_FILL

    for col_idx in range(1, col_count + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill


def _write_summary_sheet(wb: Workbook, results: List[TestResult]):
    """写入统计摘要 Sheet"""
    ws = wb.create_sheet(title="统计摘要")

    total = len(results)
    success_count = sum(1 for r in results if r.success)
    pass_count = sum(1 for r in results if r.passed)
    fail_count = sum(1 for r in results if r.success and not r.passed)
    error_count = sum(1 for r in results if r.error)

    summary_data = [
        ("指标", "数量", "占比"),
        ("总测试数", total, "100%"),
        ("请求成功", success_count, f"{success_count / total * 100:.1f}%" if total else "0%"),
        ("校验通过", pass_count, f"{pass_count / total * 100:.1f}%" if total else "0%"),
        ("校验不通过", fail_count, f"{fail_count / total * 100:.1f}%" if total else "0%"),
        ("请求异常", error_count, f"{error_count / total * 100:.1f}%" if total else "0%"),
    ]

    for row_idx, (label, value, pct) in enumerate(summary_data, start=1):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=(row_idx == 1))
        ws.cell(row=row_idx, column=2, value=value).font = Font(bold=(row_idx == 1))
        ws.cell(row=row_idx, column=3, value=pct).font = Font(bold=(row_idx == 1))

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12

    # 按 itemId 分组统计
    item_stats = {}
    for r in results:
        if r.item_id not in item_stats:
            item_stats[r.item_id] = {"total": 0, "pass": 0, "fail": 0, "error": 0}
        item_stats[r.item_id]["total"] += 1
        if r.passed:
            item_stats[r.item_id]["pass"] += 1
        elif r.error:
            item_stats[r.item_id]["error"] += 1
        else:
            item_stats[r.item_id]["fail"] += 1

    start_row = len(summary_data) + 3
    ws.cell(row=start_row, column=1, value="按模板分组统计").font = Font(bold=True, size=12)

    headers = ["itemId", "总数", "通过", "不通过", "异常"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=start_row + 1, column=col_idx, value=h).font = Font(bold=True)

    for i, (item_id, stats) in enumerate(sorted(item_stats.items()), start=start_row + 2):
        ws.cell(row=i, column=1, value=item_id)
        ws.cell(row=i, column=2, value=stats["total"])
        ws.cell(row=i, column=3, value=stats["pass"])
        ws.cell(row=i, column=4, value=stats["fail"])
        ws.cell(row=i, column=5, value=stats["error"])
