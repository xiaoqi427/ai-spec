from __future__ import annotations

import os
from typing import List

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

from .case_model import TestCase


def _get_column_index_by_name(header_row, column_name: str) -> int:
    """根据列名在表头中找到列索引（1-based）。如不存在则抛出 KeyError。"""

    normalized = column_name.strip().lower()
    for cell in header_row:
        if isinstance(cell.value, str) and cell.value.strip().lower() == normalized:
            return cell.column
    raise KeyError(f"在 Excel 表头中未找到列名: {column_name}")


def read_test_cases(
    excel_path: str,
    sheet_name: str,
    case_id_column: str,
    description_column: str,
    system_url_column: str,
    extra_context_column: str | None = None,
) -> List[TestCase]:
    """从 Excel 读取测试用例并转换为 TestCase 列表。

    约定第一行为表头，其余行为数据。
    """

    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Excel 中不存在工作表: {sheet_name}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows())
    if not rows:
        return []

    header = rows[0]
    case_id_idx = _get_column_index_by_name(header, case_id_column)
    description_idx = _get_column_index_by_name(header, description_column)
    system_url_idx = _get_column_index_by_name(header, system_url_column)
    extra_context_idx = None
    if extra_context_column:
        try:
            extra_context_idx = _get_column_index_by_name(header, extra_context_column)
        except KeyError:
            extra_context_idx = None

    test_cases: List[TestCase] = []

    for row in rows[1:]:
        case_id = str(row[case_id_idx - 1].value).strip() if row[case_id_idx - 1].value is not None else ""
        if not case_id:
            # 没有用例 ID 的行视为无效，跳过
            continue

        description = row[description_idx - 1].value or ""
        system_url = row[system_url_idx - 1].value or ""
        extra_context = None
        if extra_context_idx is not None:
            extra_context = row[extra_context_idx - 1].value

        test_case = TestCase(
            case_id=case_id,
            description=str(description),
            system_url=str(system_url),
            extra_context=str(extra_context) if extra_context is not None else None,
        )
        test_cases.append(test_case)

    return test_cases


def write_screenshot_path(
    excel_path: str,
    sheet_name: str,
    case_id: str,
    screenshot_path: str,
    case_id_column: str,
    screenshot_column: str,
) -> None:
    """根据 case_id 将截图直接嵌入到指定列。"""

    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Excel 中不存在工作表: {sheet_name}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows())
    if not rows:
        return

    header = rows[0]
    case_id_idx = _get_column_index_by_name(header, case_id_column)
    screenshot_idx = _get_column_index_by_name(header, screenshot_column)

    target_row_num = None
    normalized_case_id = str(case_id).strip()

    for idx, row in enumerate(rows[1:], start=2):  # 从第2行开始（Excel行号）
        value = row[case_id_idx - 1].value
        if value is None:
            continue
        if str(value).strip() == normalized_case_id:
            target_row_num = idx
            break

    if target_row_num is None:
        raise ValueError(f"未在 Excel 中找到指定 case_id: {case_id}")

    # 检查图片文件是否存在
    if not os.path.exists(screenshot_path):
        print(f"警告: 图片文件不存在 {screenshot_path}")
        return

    # 获取目标单元格的列字母（如 A, B, C...）
    from openpyxl.utils import get_column_letter
    target_col_letter = get_column_letter(screenshot_idx)
    target_cell = f"{target_col_letter}{target_row_num}"

    # 删除该位置的旧图片
    images_to_remove = []
    for img in ws._images:
        if img.anchor._from.row == target_row_num - 1:  # openpyxl行索引从0开始
            if img.anchor._from.col == screenshot_idx - 1:  # 列索引也从0开始
                images_to_remove.append(img)
    
    for img in images_to_remove:
        ws._images.remove(img)
        print(f"已删除旧图片，位置: {target_cell}")

    # 设置列宽和行高以容纳图片（使用合理的尺寸避免过大）
    ws.column_dimensions[target_col_letter].width = 30  # 约200像素宽
    ws.row_dimensions[target_row_num].height = 120  # 约160像素高

    # 创建图片对象并设置大小
    img = XLImage(screenshot_path)
    # 限制图片大小（宽度最大180像素，高度最大150像素）
    original_width = img.width
    original_height = img.height
    max_width = 180
    max_height = 150
    
    # 等比例缩放
    if original_width > max_width or original_height > max_height:
        ratio = min(max_width / original_width, max_height / original_height)
        img.width = int(original_width * ratio)
        img.height = int(original_height * ratio)

    # 将图片锚定到单元格，并设置为移动和调整大小都跟随单元格
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    
    # 创建单元格锚点，图片会填充在单元格内
    marker = AnchorMarker(col=screenshot_idx-1, colOff=0, row=target_row_num-1, rowOff=0)
    size = XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height))
    img.anchor = OneCellAnchor(_from=marker, ext=size)
    
    ws.add_image(img)
    
    print(f"图片已插入到 {target_cell} (尺寸: {img.width}x{img.height})")
    wb.save(excel_path)
