#!/usr/bin/env python3
"""
测试用例 Excel 解析器
解析项目 Excel 测试用例文件，生成结构化 JSON 索引。

用法:
    python3 parse_test_cases.py --input <目录或文件> --output <输出路径>
    python3 parse_test_cases.py --input docs/销售/ --output yili-out/test-case-index/test-case-index.json
    python3 parse_test_cases.py --input docs/销售/xxx.xlsx --output yili-out/test-case-index/test-case-index.json --incremental

@author: sevenxiao
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("错误: 需要 openpyxl 库。请执行 pip install openpyxl")
    sys.exit(1)


# 标准列头映射（支持模糊匹配）
COLUMN_MAPPING = {
    "序号": "id",
    "子功能名称": "feature_name",
    "子功能名称(界面)": "feature_name",
    "用例描述": "description",
    "前置条件": "precondition",
    "操作步骤": "steps",
    "预期结果": "expected",
    "实际输入": "actual_input",
    "实际输出": "actual_output",
    "测试人": "tester",
    "测试时间": "test_time",
    "测试结论": "result",
    "问题编号": "bug_id",
    "测试截图|数据": "screenshot",
    "测试截图": "screenshot",
}

# 模板号正则 (T001-T999)
TEMPLATE_PATTERN = re.compile(r"T\d{3}")

# 功能关键词（用于生成 keyword_index）
FEATURE_KEYWORDS = [
    "查询", "新增", "删除", "修改", "导出", "导入", "保存", "提交",
    "重置", "分页", "下拉", "报账单", "发票", "收款", "付款",
    "审批", "退回", "撤回", "复制", "打印", "校验", "计算",
]


def detect_columns(header_row):
    """
    自动检测列头，返回列索引到字段名的映射。
    """
    col_map = {}
    for idx, cell in enumerate(header_row):
        val = str(cell.value or "").strip()
        if val in COLUMN_MAPPING:
            col_map[idx] = COLUMN_MAPPING[val]
        else:
            # 模糊匹配
            for key, field in COLUMN_MAPPING.items():
                if key in val or val in key:
                    col_map[idx] = field
                    break
    return col_map


def extract_keywords(text):
    """
    从文本中提取关键词（模板号 + 功能词）。
    """
    if not text:
        return []
    keywords = set()
    # 提取模板号
    templates = TEMPLATE_PATTERN.findall(text)
    keywords.update(templates)
    # 提取功能关键词
    for kw in FEATURE_KEYWORDS:
        if kw in text:
            keywords.add(kw)
    return sorted(keywords)


def extract_module_name(filename):
    """
    从文件名中提取模块名。
    例: "系统运维-基础数据配置-OTC销售-手工发票报账单配置.xlsx" → "OTC-手工发票报账单配置"
    """
    stem = Path(filename).stem
    parts = stem.split("-")
    # 尝试找到 OTC/PTP/TR 等业务域标识
    module_parts = []
    found_domain = False
    for part in parts:
        part = part.strip()
        if any(domain in part for domain in ["OTC", "PTP", "TR", "EER", "FA", "RTR"]):
            # 提取业务域前缀
            for domain in ["OTC", "PTP", "TR", "EER", "FA", "RTR"]:
                if domain in part:
                    module_parts.append(domain)
                    # 提取业务域后面的部分
                    suffix = part.replace(domain, "").strip("销售采购报销")
                    break
            found_domain = True
        elif found_domain:
            module_parts.append(part)

    if module_parts:
        return "-".join(module_parts)
    # 降级：取最后两段
    if len(parts) >= 2:
        return "-".join(parts[-2:]).strip()
    return stem


def extract_business_domain(filepath):
    """
    从文件路径提取业务域。
    例: "doc/otc/xxx.xlsx" → "otc"
    例: "docs/销售/xxx.xlsx" → "销售"
    """
    parts = Path(filepath).parts
    # 优先匹配英文业务域目录名
    domain_dirs = {"otc", "ptp", "tr", "eer", "fa", "rtr"}
    for part in parts:
        if part.lower() in domain_dirs:
            return part.lower()
    # 兼容中文目录名
    cn_domains = {"销售", "采购", "报销", "资产", "差旅", "付款", "税务"}
    for part in parts:
        if part in cn_domains:
            return part
    # 从父目录名取
    parent = Path(filepath).parent.name
    return parent if parent not in ("docs", "doc") else "未分类"


def parse_excel_file(filepath, base_dir=""):
    """
    解析单个 Excel 文件，返回模块数据。
    """
    rel_path = os.path.relpath(filepath, base_dir) if base_dir else filepath
    wb = openpyxl.load_workbook(filepath, data_only=True)
    module_name = extract_module_name(os.path.basename(filepath))
    business_domain = extract_business_domain(filepath)
    file_mtime = os.path.getmtime(filepath)
    file_mtime_str = datetime.fromtimestamp(file_mtime).isoformat()

    module_data = {
        "source_file": rel_path,
        "business_domain": business_domain,
        "total_cases": 0,
        "features": {},
    }

    all_cases = []
    all_keywords = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row < 2:
            continue

        # 检测列头（第一行）
        header_row = list(ws.iter_rows(min_row=1, max_row=1, max_col=ws.max_column))[0]
        col_map = detect_columns(header_row)

        if not col_map:
            continue

        current_feature = "未分类"

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            row_data = {}
            for idx, cell in enumerate(row):
                if idx in col_map:
                    val = cell.value
                    # 清理 Excel 错误值
                    if isinstance(val, str) and val.startswith("#"):
                        val = None
                    # 时间格式化
                    if col_map[idx] == "test_time" and val:
                        try:
                            if hasattr(val, "isoformat"):
                                val = val.strftime("%Y-%m-%d")
                            else:
                                val = str(val)
                        except Exception:
                            val = str(val)
                    row_data[col_map[idx]] = val

            # 跳过完全空行
            if not any(v for k, v in row_data.items() if k not in ("id",)):
                continue

            # 更新当前功能分组
            feature = row_data.get("feature_name")
            if feature and str(feature).strip():
                current_feature = str(feature).strip()
                # 清理功能名（去掉模块前缀）
                for prefix in ["报账单配置 - ", "收款方法 - ", "报账单配置 -", "收款方法 -"]:
                    if current_feature.startswith(prefix):
                        current_feature = current_feature[len(prefix):].strip()
                        break

            # 构造测试用例
            description = str(row_data.get("description") or "").strip()
            if not description:
                continue

            steps = str(row_data.get("steps") or "").strip()
            expected = str(row_data.get("expected") or "").strip()

            # 提取关键词
            combined_text = f"{description} {steps} {expected}"
            keywords = extract_keywords(combined_text)
            all_keywords.update(keywords)

            case = {
                "id": row_data.get("id"),
                "description": description,
                "precondition": str(row_data.get("precondition") or "").strip() or None,
                "steps": steps or None,
                "expected": expected or None,
                "actual_input": str(row_data.get("actual_input") or "").strip() or None,
                "actual_output": str(row_data.get("actual_output") or "").strip() or None,
                "result": str(row_data.get("result") or "").strip() or None,
                "bug_id": row_data.get("bug_id"),
                "keywords": keywords,
            }

            # 归类到 feature
            if current_feature not in module_data["features"]:
                module_data["features"][current_feature] = []
            module_data["features"][current_feature].append(case)
            all_cases.append(case)

    module_data["total_cases"] = len(all_cases)

    file_info = {
        "path": rel_path,
        "last_modified": file_mtime_str,
        "sheet_count": len(wb.sheetnames),
        "case_count": len(all_cases),
    }

    return module_name, module_data, file_info, all_keywords, all_cases


def build_keyword_index(modules):
    """
    构建关键词索引。
    """
    keyword_index = {}
    for module_name, module_data in modules.items():
        for feature_name, cases in module_data["features"].items():
            for case in cases:
                for kw in case.get("keywords", []):
                    if kw not in keyword_index:
                        keyword_index[kw] = []
                    if module_name not in keyword_index[kw]:
                        keyword_index[kw].append(module_name)
    return keyword_index


def build_bug_index(modules):
    """
    构建 Bug ID 索引（通过问题编号列关联）。
    """
    bug_index = {}
    for module_name, module_data in modules.items():
        for feature_name, cases in module_data["features"].items():
            for case in cases:
                bug_id = case.get("bug_id")
                if bug_id:
                    bug_id_str = str(bug_id).strip().lstrip("#")
                    if bug_id_str and bug_id_str != "None":
                        ref = f"{module_name}/{feature_name}/{case.get('id', '?')}"
                        if bug_id_str not in bug_index:
                            bug_index[bug_id_str] = []
                        bug_index[bug_id_str].append(ref)
    return bug_index


def find_excel_files(input_path):
    """
    查找所有 Excel 文件。
    """
    input_path = Path(input_path)
    if input_path.is_file() and input_path.suffix in (".xlsx", ".xls"):
        return [str(input_path)]
    elif input_path.is_dir():
        files = []
        for f in input_path.rglob("*.xlsx"):
            # 跳过临时文件
            if f.name.startswith("~$"):
                continue
            files.append(str(f))
        return sorted(files)
    else:
        print(f"警告: {input_path} 不是有效的文件或目录")
        return []


def load_existing_index(output_path):
    """
    加载已有索引（用于增量更新）。
    """
    if os.path.exists(output_path):
        with open(output_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def main():
    parser = argparse.ArgumentParser(description="解析 Excel 测试用例，生成 JSON 索引")
    parser.add_argument("--input", required=True, help="Excel 文件或目录路径")
    parser.add_argument("--output", default="yili-out/test-case-index/test-case-index.json",
                        help="输出 JSON 索引路径")
    parser.add_argument("--incremental", action="store_true",
                        help="增量更新（仅解析新增/修改的文件）")
    args = parser.parse_args()

    input_path = args.input
    output_path = args.output

    # 查找 Excel 文件
    excel_files = find_excel_files(input_path)
    if not excel_files:
        print("未找到 Excel 文件")
        sys.exit(1)

    print(f"找到 {len(excel_files)} 个 Excel 文件")

    # 增量模式：加载已有索引
    existing_index = None
    if args.incremental:
        existing_index = load_existing_index(output_path)
        if existing_index:
            existing_files = {f["path"] for f in existing_index.get("source_files", [])}
            print(f"已有索引包含 {len(existing_files)} 个文件")

    # 确定 base_dir
    base_dir = str(Path(input_path)) if Path(input_path).is_dir() else str(Path(input_path).parent)

    # 解析文件
    modules = {}
    source_files = []
    total_cases = 0

    for filepath in excel_files:
        rel = os.path.relpath(filepath, base_dir)

        # 增量模式：检查文件是否需要重新解析
        if args.incremental and existing_index:
            existing_file_info = next(
                (f for f in existing_index.get("source_files", []) if f["path"] == rel),
                None
            )
            if existing_file_info:
                file_mtime = datetime.fromtimestamp(os.path.getmtime(filepath)).isoformat()
                if existing_file_info.get("last_modified") == file_mtime:
                    print(f"  跳过（未修改）: {rel}")
                    continue

        print(f"  解析: {rel}")
        try:
            module_name, module_data, file_info, keywords, cases = parse_excel_file(filepath, base_dir)
            modules[module_name] = module_data
            source_files.append(file_info)
            total_cases += len(cases)
            print(f"    → 模块: {module_name}, 用例: {len(cases)} 条")
        except Exception as e:
            print(f"    × 解析失败: {e}")

    # 增量模式：合并已有数据
    if args.incremental and existing_index:
        for mod_name, mod_data in existing_index.get("modules", {}).items():
            if mod_name not in modules:
                modules[mod_name] = mod_data
        for sf in existing_index.get("source_files", []):
            if not any(f["path"] == sf["path"] for f in source_files):
                source_files.append(sf)

    # 构建索引
    keyword_index = build_keyword_index(modules)
    bug_index = build_bug_index(modules)

    # 组装最终结构
    index = {
        "version": "1.0",
        "parsed_at": datetime.now().isoformat(),
        "source_dir": base_dir,
        "source_files": source_files,
        "total_cases": sum(m["total_cases"] for m in modules.values()),
        "modules": modules,
        "keyword_index": keyword_index,
        "bug_index": bug_index,
    }

    # 确保输出目录存在
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # 写入 JSON
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)

    print(f"\n索引生成完毕:")
    print(f"  文件数: {len(source_files)}")
    print(f"  模块数: {len(modules)}")
    print(f"  用例总数: {index['total_cases']}")
    print(f"  关键词数: {len(keyword_index)}")
    print(f"  Bug 关联数: {len(bug_index)}")
    print(f"  输出: {output_path}")


if __name__ == "__main__":
    main()
